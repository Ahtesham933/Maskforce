import base64
import json
import os
from datetime import datetime
from io import BytesIO

import cv2
import numpy as np
from django.core.files.base import ContentFile
from django.http import JsonResponse, HttpResponse
from django.shortcuts import render
from django.views.decorators.csrf import csrf_exempt
from django.db.models import Count, Q
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage

from .models import DetectionLog
from .ml.detector import detect_from_frame


def index(request):
    return render(request, 'index.html')


@csrf_exempt
def detect_api(request):
    """
    Receives a base64-encoded JPEG frame from the browser,
    runs face + mask detection, and -- only if the mask status
    has changed since the last saved record -- saves a new
    DetectionLog entry (with cropped face image).
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)

    try:
        data = json.loads(request.body)
        image_data = data.get('image', '')

        # strip "data:image/jpeg;base64," prefix if present
        if ',' in image_data:
            image_data = image_data.split(',', 1)[1]

        img_bytes = base64.b64decode(image_data)
        np_arr = np.frombuffer(img_bytes, np.uint8)
        frame = cv2.imdecode(np_arr, cv2.IMREAD_COLOR)

        if frame is None:
            return JsonResponse({'error': 'Invalid image'}, status=400)

    except Exception as e:
        return JsonResponse({'error': f'Failed to decode image: {e}'}, status=400)

    detections = detect_from_frame(frame)

    if not detections:
        return JsonResponse({
            'face_detected': False,
            'mask': None,
            'confidence': None,
            'saved': False,
            'box': None,
            'frame_width': frame.shape[1],
            'frame_height': frame.shape[0],
        })

    # use the first (largest-confidence) detected face
    result = detections[0]
    mask_detected = result['mask_detected']
    confidence = result['confidence']
    sx, sy, ex, ey = result['box']

    # check the last saved record's status
    last_log = DetectionLog.objects.order_by('-timestamp').first()
    status_changed = (last_log is None) or (last_log.mask_detected != mask_detected)

    saved = False
    if status_changed:
        # encode the cropped face as jpg
        success, buffer = cv2.imencode('.jpg', result['face_crop'])
        if success:
            filename = f"{datetime.now().strftime('%Y%m%d_%H%M%S_%f')}.jpg"
            log = DetectionLog(
                mask_detected=mask_detected,
                confidence=confidence,
            )
            log.image_path.save(filename, ContentFile(buffer.tobytes()), save=True)
            saved = True

    return JsonResponse({
        'face_detected': True,
        'mask': mask_detected,
        'confidence': confidence,
        'saved': saved,
        'box': {'x1': int(sx), 'y1': int(sy), 'x2': int(ex), 'y2': int(ey)},
        'frame_width': frame.shape[1],
        'frame_height': frame.shape[0],
    })


def logs_api(request):
    """Returns all detection logs as JSON, most recent first."""
    logs = DetectionLog.objects.all()[:200]  # cap to last 200 records

    data = []
    for log in logs:
        data.append({
            'id': log.id,
            'mask_detected': log.mask_detected,
            'confidence': log.confidence,
            'timestamp': log.timestamp.strftime('%Y-%m-%d %H:%M:%S'),
            'image_url': log.image_path.url if log.image_path else None,
        })

    return JsonResponse({'logs': data})


def stats_api(request):
    """Returns aggregated stats for the dashboard charts."""
    total = DetectionLog.objects.count()
    mask_count = DetectionLog.objects.filter(mask_detected=True).count()
    no_mask_count = total - mask_count

    mask_pct = round((mask_count / total) * 100, 1) if total else 0
    no_mask_pct = round((no_mask_count / total) * 100, 1) if total else 0

    # daily breakdown for the chart -- grouped in Python to avoid
    # SQLite/TruncDate timezone quirks that can return NULL dates
    daily_map = {}
    for log in DetectionLog.objects.all().order_by('timestamp'):
        local_ts = timezone.localtime(log.timestamp) if timezone.is_aware(log.timestamp) else log.timestamp
        date_key = local_ts.strftime('%Y-%m-%d')

        if date_key not in daily_map:
            daily_map[date_key] = {'mask': 0, 'no_mask': 0}

        if log.mask_detected:
            daily_map[date_key]['mask'] += 1
        else:
            daily_map[date_key]['no_mask'] += 1

    daily_data = [
        {'date': date_key, 'mask': counts['mask'], 'no_mask': counts['no_mask']}
        for date_key, counts in sorted(daily_map.items())
    ]

    return JsonResponse({
        'total': total,
        'mask_count': mask_count,
        'no_mask_count': no_mask_count,
        'mask_pct': mask_pct,
        'no_mask_pct': no_mask_pct,
        'daily': daily_data,
    })


@csrf_exempt
def delete_logs_api(request):
    """
    Receives a list of log IDs in the request body and deletes
    those DetectionLog records, along with their saved image files.
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)

    try:
        data = json.loads(request.body)
        ids = data.get('ids', [])
    except Exception as e:
        return JsonResponse({'error': f'Invalid request: {e}'}, status=400)

    if not ids:
        return JsonResponse({'error': 'No ids provided'}, status=400)

    logs = DetectionLog.objects.filter(id__in=ids)
    deleted_count = 0

    for log in logs:
        # remove the image file from disk if it exists
        if log.image_path:
            log.image_path.delete(save=False)
        log.delete()
        deleted_count += 1

    return JsonResponse({'deleted': deleted_count})


def export_logs_api(request):
    """Generates an .xlsx file of all detection logs, with embedded
    60x60px thumbnail images, and returns it for download."""
    logs = DetectionLog.objects.all().order_by('-timestamp')

    wb = Workbook()
    ws = wb.active
    ws.title = "Detection Logs"

    THUMB_SIZE = 60  # px

    headers = ["Snapshot", "ID", "Status", "Confidence (%)", "Timestamp"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # column A holds the image; give it enough width for a 60px thumbnail
    ws.column_dimensions['A'].width = 10

    row_idx = 2
    for log in logs:
        status = "Mask" if log.mask_detected else "No Mask"
        confidence_pct = round(log.confidence * 100, 1)
        timestamp_str = log.timestamp.strftime('%Y-%m-%d %H:%M:%S')

        ws.cell(row=row_idx, column=2, value=log.id)
        ws.cell(row=row_idx, column=3, value=status)
        ws.cell(row=row_idx, column=4, value=confidence_pct)
        ws.cell(row=row_idx, column=5, value=timestamp_str)

        # embed a resized thumbnail if the image file exists on disk
        if log.image_path and os.path.exists(log.image_path.path):
            try:
                pil_img = PILImage.open(log.image_path.path)
                pil_img = pil_img.convert('RGB')
                pil_img.thumbnail((THUMB_SIZE, THUMB_SIZE))

                img_buffer = BytesIO()
                pil_img.save(img_buffer, format='PNG')
                img_buffer.seek(0)

                xl_img = XLImage(img_buffer)
                xl_img.width = THUMB_SIZE
                xl_img.height = THUMB_SIZE
                ws.add_image(xl_img, f'A{row_idx}')
            except Exception:
                pass  # skip embedding if the image can't be read/processed

        # set row height so the thumbnail fits (Excel row height units ≈ px * 0.75)
        ws.row_dimensions[row_idx].height = THUMB_SIZE * 0.75

        row_idx += 1

    # auto-size the non-image columns based on content length
    for col_idx in range(2, 6):
        col_letter = get_column_letter(col_idx)
        max_length = max(
            (len(str(ws.cell(row=r, column=col_idx).value))
             for r in range(1, row_idx) if ws.cell(row=r, column=col_idx).value is not None),
            default=8
        )
        ws.column_dimensions[col_letter].width = max_length + 4

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"mask_detection_logs_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response